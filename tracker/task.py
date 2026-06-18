import os
import requests
import pandas as pd
from decimal import Decimal

from django.conf import settings
from django.utils import timezone
from celery import shared_task

from .models import Asset, AssetPriceHistory, Transaction


@shared_task
def fetch_daily_crypto_prices():
    crypto_assets = Asset.objects.filter(asset_type='CRYPTO')
    if not crypto_assets.exists():
        return "Немає криптоактивів для оновлення"

    base_url = "https://api.binance.com/api/v3/ticker/price"
    today = timezone.now().date()
    updated_count = 0

    for asset in crypto_assets:
        symbol = f"{asset.ticker.upper()}USDT"

        try:
            response = requests.get(base_url, params={'symbol': symbol}, timeout=5)
            data = response.json()

            if 'price' in data:
                price = Decimal(data['price'])

                AssetPriceHistory.objects.update_or_create(
                    asset=asset,
                    date=today,
                    defaults={'close_price': price}
                )
                updated_count += 1
        except Exception as e:
            print(f"Помилка оновлення ціни для {asset.ticker}: {str(e)}")

    return f"Успішно оновлено цін: {updated_count}"


@shared_task
def generate_excel_report(user_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    transactions = Transaction.objects.filter(user_id=user_id).values(
        'timestamp', 'asset__ticker', 'transaction_type', 'quantity', 'execution_price'
    ).order_by('-timestamp')

    if not transactions:
        return None

    df = pd.DataFrame.from_records(transactions)

    df.rename(columns={
        'timestamp': 'Дата та час',
        'asset__ticker': 'Актив',
        'transaction_type': 'Тип операції',
        'quantity': 'Кількість',
        'execution_price': 'Ціна виконання ($)'
    }, inplace=True)

    df['Кількість'] = df['Кількість'].astype(float)
    df['Ціна виконання ($)'] = df['Ціна виконання ($)'].astype(float)
    df['Загальна сума ($)'] = (df['Кількість'] * df['Ціна виконання ($)']).round(2)

    df['Дата та час'] = df['Дата та час'].dt.tz_localize(None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Транзакції"

    ws["A1"] = "Фінансовий звіт FinTracker"
    ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    ws.row_dimensions[1].height = 25

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 3:
                cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")

                if c_idx == 1:
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
                elif c_idx == 2:
                    cell.font = Font(bold=True)
                elif c_idx == 3:
                    if value == "BUY":
                        cell.font = Font(bold=True, color="008000")
                    else:
                        cell.font = Font(bold=True, color="FF0000")
                elif c_idx == 4:
                    cell.number_format = "#,##0.0000"
                elif c_idx in [5, 6]:
                    cell.number_format = "$#,##0.00"

            thin = Side(border_style="thin", color="D9D9D9")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"finance_report_user_{user_id}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)

    return f"/media/reports/{filename}"